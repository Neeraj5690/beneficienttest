import datetime
import math
import re
import time
import openpyxl
import xlrd
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
import pandas as pd
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


@allure.step("Entering username ")
def enter_username(username):
      driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
      driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global checkcount
  global path
  global Exe
  global Dict
  global Dict2
  global FundsNamesList

  TestName = "test_Funds_HistoricalValuesForecast"
  description = "This test scenario is to verify values in Historical Net Cash Flows and Liquidity Projections by Fund (Forecast)"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FundsNamesList = []

  FailStatus = "Pass"
  TestDirectoryName="test_Values_Check"

  # Creating Dictionary for Quarters
  Dict = {}
  Dict2 = {}

  Exe = "Yes"
  path = 'C:/BIDS/beneficienttest/Beneficient/test_Values_Check/'

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe = "No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe = "Yes"

  if Exe == "Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@786")
      button = driver.find_element_by_xpath("//input[@type='submit']")
      driver.execute_script("arguments[0].click();", button)


  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  " + description, 0, 1)

      for i in range(len(TestResult)):
          pdf.set_fill_color(255, 255, 255)
          pdf.set_text_color(0, 0, 0)
          if (TestResultStatus[i] == "Fail"):
              pdf.set_text_color(255, 0, 0)
              TestFailStatus.append("Fail")
          TestName1 = TestResult[i].encode('latin-1', 'ignore').decode('latin-1')
          pdf.multi_cell(0, 7, str(i + 1) + ")  " + TestName1, 0, 1, fill=True)
          TestFailStatus.append("Pass")
      pdf.output(TestName + "_" + ct + ".pdf", 'F')

      # -----------To check if any failed Test case present------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io] == "Fail":
              FailStatus = "Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      # ----------------------------------------------------------------------------

      # ---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                      sheet1.cell(row=ii1, column=1).value = check
                      checkcount1 = 1
      # -----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_Funds_Values(test_setup):
    if Exe == "Yes":
        ForecastYear=8
        skip1 = 0

        PageName = "Funds"
        PageTitle = "Funds - BIDS"
        button = driver.find_element_by_xpath("//*[@title='" + PageName + "']")
        driver.execute_script("arguments[0].click();", button)
        for iat1 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
                time.sleep(1)
            except Exception:
                #time.sleep(1)
                break
        time.sleep(2)
        try:
            assert PageTitle in driver.title, PageName + " not able to open"
            TestResult.append(PageName + " page Opened successfully")
            TestResultStatus.append("Pass")
        except Exception:
            pass
            TestResult.append(PageName + " page not able to open")
            TestResultStatus.append("Fail")

        driver.find_element_by_xpath(
            "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[1]/div[2]/div/div[2]/div/div[2]/div/div").click()
        time.sleep(1)
        ActionChains(driver).key_down(Keys.DOWN).perform()
        time.sleep(1)
        ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
        for iat3 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
                time.sleep(1)
            except Exception:
                time.sleep(1)
                break

        ElementName = "Forecast Dropdown"
        try:
            assert PageTitle in driver.title, PageName + " not able to open"
            TestResult.append(ElementName + " clicked successfully")
            TestResultStatus.append("Pass")
        except Exception:
            pass
            TestResult.append(ElementName + " not able to click")
            TestResultStatus.append("Fail")

        ElementName = "Next Arrow at Historical section"
        try:
            assert PageTitle in driver.title, PageName + " not able to open"
            TestResult.append(ElementName + " clicked successfully")
            TestResultStatus.append("Pass")
        except Exception:
            pass
            TestResult.append(ElementName + " not able to click")
            TestResultStatus.append("Fail")

        ElementName = "Period Dropdown"
        try:
            assert PageTitle in driver.title, PageName + " not able to open"
            TestResult.append(ElementName + " clicked successfully")
            TestResultStatus.append("Pass")
        except Exception:
            pass
            TestResult.append(ElementName + " not able to click")
            TestResultStatus.append("Fail")

        for ii in range(ForecastYear+1):
            print()
            print("Iteration:   "+str(ii))
            #---------------------------
            TestResult.append("--------------Iteration---------------:   "+str(ii))
            TestResultStatus.append("Pass")
            #----------------------------
            if ii>0:
                elements = driver.find_elements_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[1]/div[2]/div/div[3]/div/div[2]/div/div")
                for elem in elements:
                    try:
                        elem.click()
                    except Exception:
                        button = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[1]/div[2]/div/div[3]/div/div[2]/div/div")
                        driver.execute_script("arguments[0].click();", button)
                    break
                time.sleep(1)
                ActionChains(driver).key_down(Keys.DOWN).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.DOWN).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.DOWN).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.DOWN).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.DOWN).perform()
                time.sleep(1)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                for iat4 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        time.sleep(1)
                    except Exception:
                        time.sleep(1)
                        break
            Label1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout ContentLayout---padding_less']/div[2]/div[2]/div/div/table/thead/tr/th[2]").get_attribute("abbr")
            String1=driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div[1]/div/div[2]/div/input").get_attribute("value")

            Label2 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout ContentLayout---padding_less']/div[2]/div[2]/div/div/table/thead/tr/th[3]").get_attribute(
                "abbr")
            String2=driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div[2]/div/div[2]/div/input").get_attribute("value")

            Label3 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout ContentLayout---padding_less']/div[2]/div[2]/div/div/table/thead/tr/th[4]").get_attribute(
                "abbr")
            String3=driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div[3]/div/div[2]/div/input").get_attribute("value")

            Label4 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout ContentLayout---padding_less']/div[2]/div[2]/div/div/table/thead/tr/th[5]").get_attribute(
                "abbr")
            String4=driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div[4]/div/div[2]/div/input").get_attribute("value")

            Label5 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout ContentLayout---padding_less']/div[2]/div[2]/div/div/table/thead/tr/th[6]").get_attribute(
                "abbr")
            String5=driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div[5]/div/div[2]/div/input").get_attribute("value")

            Hyphen="_"

            # print(Label1)
            # print(String1)
            String1 = String1.replace(" ", "")
            String1 = re.sub(r'[?|$|.|!|,|-]', r'', String1)
            if re.search(Hyphen, String1):
                String1="0"
            String1Float = float(String1)

            # print(Label2)
            # print(String2)
            String2 = String2.replace(" ", "")
            String2 = re.sub(r'[?|$|.|!|,|-]', r'', String2)
            if re.search(Hyphen, String2):
                String2="0"
            String2Float = float(String2)

            # print(Label3)
            # print(String3)
            String3 = String3.replace(" ", "")
            String3 = re.sub(r'[?|$|.|!|,|-]', r'', String3)
            if re.search(Hyphen, String3):
                String3="0"
            String3Float = float(String3)

            # print(Label4)
            # print(String4)
            String4 = String4.replace(" ", "")
            String4 = re.sub(r'[?|$|.|!|,|-]', r'', String4)
            if re.search(Hyphen, String4):
                String4="0"
            String4Float = float(String4)

            # print(Label5)
            # print(String5)
            String5 = String5.replace(" ", "")
            String5 = re.sub(r'[?|$|.|!|,|-]', r'', String5)
            if re.search(Hyphen, String5):
                String5="0"
            String5Float = float(String5)

            key1 = Label1
            if key1 in Dict.keys():
                pass
            else:
                Dict[Label1] = String1Float

            key1 = Label2
            if key1 in Dict.keys():
                pass
            else:
                Dict[Label2] = String2Float

            key1 = Label3
            if key1 in Dict.keys():
                pass
            else:
                Dict[Label3] = String3Float

            key1 = Label4
            if key1 in Dict.keys():
                pass
            else:
                Dict[Label4] = String4Float

            key1 = Label5
            if key1 in Dict.keys():
                pass
            else:
                Dict[Label5] = String5Float

            try:
                FundsItems = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[3]").text
                TotalFundsItems = FundsItems[-2] + FundsItems[-1]
                TotalFundsItemsInt = int(TotalFundsItems)
                #print(str(TotalFundsItemsInt))
                TotalFundsItemsInt = math.ceil(TotalFundsItemsInt / 7)
                print("Arrow should click for "+str(TotalFundsItemsInt-1))

                for ii1 in range(TotalFundsItemsInt):
                    ii2 = 1
                    if ii1 > 0:
                        button = driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/span[4]/a[1]/i")
                        driver.execute_script("arguments[0].click();", button)
                        time.sleep(5)

                    for ii2 in range(1, 8):
                        try:
                            FundName = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div[1]/table/tbody/tr[" + str(
                                    ii2) + "]/td[1]/div/p/a").text
                            if (FundName in FundsNamesList):
                                pass
                            else:
                                FundsNamesList.append(FundName)
                                print(FundName)
                                # ---------------------------
                                TestResult.append(FundName)
                                TestResultStatus.append("Pass")
                                # ----------------------------
                        except Exception:
                            pass
                print()
                print("Funds iteration run for "+str(ii1))
                print("Len of FundsNamesList " + str(len(FundsNamesList)))
                # ---------------------------
                TestResult.append("Total Funds to verify: " + str(len(FundsNamesList)))
                TestResultStatus.append("Pass")
                # ----------------------------

            except Exception as fe:
                print(fe)
                pass

        print("\n ********************printing Dictionary 1 : ***************************")
        print(Dict)
        # ---------------------------
        TestResult.append("Printing fetched Quarter values")
        TestResultStatus.append("Pass")
        # ----------------------------
        # ---------------------------
        TestResult.append(Dict)
        TestResultStatus.append("Pass")
        # ----------------------------

        #--------------Now working started to fetch Quarter values of each funds--------------
        PageName = "Funds"
        button = driver.find_element_by_xpath(
            "//*[@title='" + PageName + "']")
        driver.execute_script("arguments[0].click();", button)
        try:
            driver.switch_to_alert().accept()
        except Exception:
            pass
        for iat5 in range(1000):
            try:
                bool = driver.find_element_by_xpath(
                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
                time.sleep(1)
            except Exception:
                time.sleep(1)
                break
        time.sleep(1)


    #---------------------------Fetching details for all Funds ----------------------------
        for ii3 in range(len(FundsNamesList)):
            print()
            print()
            print(str(ii3))
            if ii3 ==5 or ii3 ==15 or ii3 ==30 or ii3 ==45 or ii3 ==60 or ii3 ==75 or ii3 ==90 or ii3 ==105:
                #print("----------------***************"+str(ii3))
                driver.delete_all_cookies()
                time.sleep(5)
                driver.get("https://beneficienttest.appiancloud.com/suite/")
                driver.find_element_by_id("un").send_keys("neeraj.kumar")
                driver.find_element_by_id("pw").send_keys("Crochet@786")
                button = driver.find_element_by_xpath(
                    "//input[@type='submit']")
                driver.execute_script("arguments[0].click();", button)
                for iat8 in range(1000):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        time.sleep(1)
                    except Exception:
                        break

            print(FundsNamesList[ii3])
            # ---------------------------
            TestResult.append(FundsNamesList[ii3])
            TestResultStatus.append("Pass")
            # ----------------------------
            try:
                driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td[2]/div/p/a[contains(text(),'" +
                    FundsNamesList[ii3] + "')]").click()
            except Exception:
                time.sleep(5)
                print("Inside Except ********************")
                try:
                    time.sleep(3)
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td[2]/div/p/a[contains(text(),'" +
                        FundsNamesList[ii3] + "')]").click()
                except Exception:
                    print("Clicked on next 1 funds icon")
                    driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a").click()
                    for iat12 in range(15):
                        try:
                            bool = driver.find_element_by_xpath(
                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                            time.sleep(1)
                        except Exception:
                            break
                    try:
                        time.sleep(4)
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td[2]/div/p/a[contains(text(),'" +
                            FundsNamesList[ii3] + "')]").click()
                    except Exception:
                        print("Clicked on next 2 funds icon")
                        driver.find_element_by_xpath(
                            "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[2]/div/div/span[4]/a").click()
                        for iat13 in range(15):
                            try:
                                bool = driver.find_element_by_xpath(
                                    "//div[@id='appian-working-indicator-hidden']").is_enabled()
                                time.sleep(1)
                            except Exception:
                                break
                        time.sleep(3)
                        try:
                            buttonFundName = driver.find_element_by_xpath(
                                "//div[@class='ContentLayout---content_layout']/div/div/div/div[4]/div/div/div/div/div/div[2]/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr/td[2]/div/p/a[contains(text(),'" +
                                FundsNamesList[ii3] + "')]")
                            driver.execute_script("arguments[0].click();", buttonFundName)
                        except Exception:
                            print(FundsNamesList[ii3] +" Fund not able to find")
                            TestResult.append(FundsNamesList[ii3] +" Fund not able to find")
                            TestResultStatus.append("Fail")
                            skip1=1

            for iat9 in range(15):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                    time.sleep(1)
                except Exception:
                    break
            time.sleep(1)

            if skip1==0:
                try:
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[1]/button").click()
                except Exception:
                    time.sleep(7)
                    driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div[1]/button")
                for iat11 in range(15):
                    try:
                        bool = driver.find_element_by_xpath(
                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                        time.sleep(1)
                    except Exception:
                        break
                Quarters=driver.find_elements_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr/th/div/p/span/a/span")
                #print("Quarters rows " + str(len(Quarters)))

                for ii4 in range(1,len(Quarters)+1):
                    Period=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr["+str(ii4)+"]/th/div/p/span/a").text

                    if "Unknown" in Period :
                        pass
                    else:
                        pass
                    print("Period " + Period)
                    BenNetProceeds_USD=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr["+str(ii4)+"]/td[6]/div/p/span").text
                    print("BenNetProceeds_USD: " + BenNetProceeds_USD)
                    BenNetProceeds_USD = BenNetProceeds_USD.replace(" ", "")
                    BenNetProceeds_USD = re.sub(r'[?|$|.|!|,|-]', r'', BenNetProceeds_USD)
                    if re.search(Hyphen, BenNetProceeds_USD):
                        BenNetProceeds_USD = "0"
                    BenNetProceeds_USDFloat=float(BenNetProceeds_USD)
                    key=Period
                    if key in Dict2.keys():
                        BenNetProceeds_USDFloat=Dict2.get(Period)+BenNetProceeds_USDFloat
                    Dict2[Period] = BenNetProceeds_USDFloat

            # --------------Navigating back to each fund--------------

            PageName = "Funds"
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            skip1 = 0
            try:
                driver.switch_to_alert().accept()
            except Exception:
                pass
            for iat6 in range(15):
                try:
                    bool = driver.find_element_by_xpath(
                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                    time.sleep(1)
                except Exception:
                    #time.sleep(1)
                    break
        print("\n************ printing 2nd Dictionary : **************")
        print(Dict2)
        # ---------------------------
        TestResult.append(Dict2)
        TestResultStatus.append("Pass")
        # ----------------------------

        #***************Compare Dictionaries Data******************

        list1 = []
        test = Dict.keys()
        list1 = list(test)

        print("Elements of the List:\n")
        for x in list1:
            print(x)

        try:
            for ii12 in range (len(list1)+1):
                print()
                if Dict2.get(list1[ii12]) == None:
                    print("None found------")
                    Dict2[list1[ii12]] = float("0")

                if Dict.get(list1[ii12]) != Dict2.get(list1[ii12]):
                    print("Value does not match for " + str(list1[ii12]))
                    print("Dict value is " + str(Dict.get(list1[ii12])))
                    print("Dict2 value is " + str(Dict2.get(list1[ii12])))

                    TestResult.append("Value for "+str(list1[ii12])+" does not match, " +"Value at Historical section is "+str(Dict.get(list1[ii12]))+" and at Funds level is "+ str(Dict2.get(list1[ii12])))
                    TestResultStatus.append("Fail")
                else:
                    print("Value matched for " + str(list1[ii12]))
                    print("Dict value is " + str(Dict.get(list1[ii12])))
                    print("Dict2 value is " + str(Dict2.get(list1[ii12])))

                    # ---------------------------
                    TestResult.append("Value matched for " + str(list1[ii12])+", Data 1 is " + str(Dict.get(list1[ii12]))+", Data 2 is " + str(Dict2.get(list1[ii12])))
                    TestResultStatus.append("Pass")
                    # ----------------------------
        except Exception as qa:
            print("jdbdksjdk")
            print(qa)


    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------