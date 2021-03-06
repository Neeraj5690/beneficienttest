import datetime
import time
import openpyxl
import xlrd
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
import pandas as pd
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

@allure.step("Entering username ")
def enter_username(username):
      driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
      driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global checkcount
  TestName = "test_VerifyingAllLinksFunds"
  description = "This is smoke test case to verify all sections inside Funds modules"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus = "Pass"
  TestDirectoryName="test_Funds"
  global Exe
  Exe = "Yes"

  ExcelFileName = "Execution"
  locx = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke/Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe = "No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe = "Yes"

  if Exe == "Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@7866")
      driver.find_element_by_xpath("//input[@type='submit']").click()


  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image('C:/BIDS/beneficienttest/Beneficient/test_Smoke/EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
      pdf.multi_cell(0, 20, "Description:  " + description, 0, 1)

      for i in range(len(TestResult)):
          pdf.set_fill_color(255, 255, 255)
          pdf.set_text_color(0, 0, 0)
          if (TestResultStatus[i] == "Fail"):
              pdf.set_text_color(255, 0, 0)
              TestFailStatus.append("Fail")
          TestName1 = TestResult[i].encode('latin-1', 'ignore').decode('latin-1')
          pdf.multi_cell(0, 7, str(i + 1) + ")  " + TestName1, 0, 1, fill=True)
          TestFailStatus.append("Pass")
      pdf.output(TestName + "_" + ct + ".pdf", 'F')

      # -----------To check if any failed Test case present------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io] == "Fail":
              FailStatus = "Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = ('C:/BIDS/beneficienttest/Beneficient/PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      # ----------------------------------------------------------------------------

      # ---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = ('C:/BIDS/beneficienttest/Beneficient/test_Smoke/Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                      sheet1.cell(row=ii1, column=1).value = check
                      checkcount1 = 1
      # -----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerfyAllLinksFundsPage(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            PageName = "Funds"
            PageTitle = "Funds - BIDS"
            loc = ("C:/BIDS/beneficienttest/Beneficient/test_Smoke/XpathDataLinks/Main.xls")

            wb = xlrd.open_workbook(loc)
            sheet = wb.sheet_by_index(0)
            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                assert PageTitle in driver.title, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)

            CheckParent = 0
            for ia in range(100):
                ia = ia + 1
                # print()
                # print("ia is " + str(ia))
                try:
                    bool_series = pd.isnull(sheet.cell_value(ia, 0))
                    # print("bool_series is "+ sheet.cell_value(ia, 0))
                    if (bool_series == True):
                        break
                    else:
                        if (sheet.cell_value(ia, 3) == "No"):
                            if (sheet.cell_value(ia, 0) == PageName):
                                print()
                                try:
                                    InOrOut = sheet.cell_value(ia, 9)
                                    #print("InOrOut is " + InOrOut)
                                    if InOrOut == "Inside":
                                        print("CheckParent 1 is " + str(CheckParent))
                                        if sheet.cell_value(ia-1, 9)=="Inside":
                                            print("Going Inside")
                                            CheckParentPre=sheet.cell_value((ia-1), 10)
                                            CheckParentCurr = sheet.cell_value((ia), 10)
                                            print("CheckParentPre is "+CheckParentPre)
                                            print("CheckParentCurr is " + CheckParentCurr)
                                            if CheckParentPre in CheckParentCurr:
                                                print("Pre parent is same")
                                                CheckParent=1
                                                try:
                                                    driver.find_element_by_xpath(sheet.cell_value(ia, 10)).click()
                                                except Exception:
                                                    print("No need to go back to parent page handled")
                                                    pass

                                        # CheckParent is 0 when coming first time inside funds
                                        if CheckParent==0:
                                            for iat10 in range(500):
                                                try:
                                                    bool = driver.find_element_by_xpath(
                                                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                                                except Exception:
                                                    time.sleep(1)
                                                    break
                                            time.sleep(2)
                                            driver.find_element_by_xpath(sheet.cell_value(ia, 10)).click()
                                            print("Parent Page link clicked ")
                                        try:
                                            WebDriverWait(driver, SHORT_TIMEOUT
                                                          ).until(
                                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                            WebDriverWait(driver, LONG_TIMEOUT
                                                          ).until_not(
                                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                        except TimeoutException:
                                            pass
                                        time.sleep(1)
                                        #--------------------------------------------------
                                        if sheet.cell_value(ia, 1)=="Investments":
                                            print("First default tab so no need to perform click to navigate")
                                            driver.refresh()
                                            try:
                                                driver.switch_to_alert().accept()
                                            except Exception:
                                                pass
                                            try:
                                                WebDriverWait(driver, SHORT_TIMEOUT
                                                              ).until(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                                WebDriverWait(driver, LONG_TIMEOUT
                                                              ).until_not(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                            except TimeoutException:
                                                pass
                                            time.sleep(1)
                                            print("Page refreshed to handle default (Investment) page")
                                        else:
                                            try:
                                                driver.find_element_by_xpath(sheet.cell_value(ia, 2)).click()
                                                print("Page clicked")
                                            except Exception:
                                                time.sleep(5)
                                                print("Waited for some time to click on the page ")
                                                driver.find_element_by_xpath(sheet.cell_value(ia, 2)).click()
                                    elif InOrOut == "Outside":
                                        print("Inside Outside")
                                        print("CheckParent 2 is "+str(CheckParent))
                                        if CheckParent==1:
                                            driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
                                            time.sleep(1)
                                            try:
                                                driver.switch_to_alert().accept()
                                            except Exception:
                                                pass
                                        try:
                                            WebDriverWait(driver, SHORT_TIMEOUT
                                                          ).until(
                                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                            WebDriverWait(driver, LONG_TIMEOUT
                                                          ).until_not(
                                                EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                        except TimeoutException:
                                            pass
                                        time.sleep(1)
                                        driver.find_element_by_xpath(sheet.cell_value(ia, 2)).click()

                                    print("Verification started for:  " + sheet.cell_value(ia, 1))
                                    for iat21 in range(1000):
                                        try:
                                            bool = driver.find_element_by_xpath(
                                                "//div[@id='appian-working-indicator-hidden']").is_enabled()
                                        except Exception:
                                            time.sleep(1)
                                            break
                                    # print("link clicked:  " + sheet.cell_value(ia, 1))
                                    # print("Skip is " + sheet.cell_value(ia, 3))
                                    DoubleClick = sheet.cell_value(ia, 4)
                                    # print("DoubleClick is "+DoubleClick)
                                    NaviBack = sheet.cell_value(ia, 5)
                                    # print("NaviBack is " + NaviBack)
                                    TitleVerify = sheet.cell_value(ia, 6)
                                    # print("TitleVerify is " + TitleVerify)
                                    TitleToVerify = sheet.cell_value(ia, 7)
                                    # print("TitleToVerify is " + TitleToVerify)
                                    TitleLink = sheet.cell_value(ia, 8)
                                    # print("TitleLink is " + TitleLink)

                                    if DoubleClick == "Yes":
                                        driver.find_element_by_xpath(sheet.cell_value(ia, 2)).click()
                                        time.sleep(1)
                                        # print("Link again clicked for  " + sheet.cell_value(ia, 1))

                                    elif DoubleClick == "No":
                                        # print("Inside Double clicked NO")
                                        if NaviBack == "Yes" and TitleVerify == "No":
                                            # print("Inside NaviBack=Yes TitleVerify= NO")
                                            for iat3 in range(1000):
                                                try:
                                                    bool = driver.find_element_by_xpath(
                                                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                                                except Exception:
                                                    time.sleep(1)
                                                    break
                                            # print("Browser Back clicked for  " + sheet.cell_value(ia, 1))
                                            time.sleep(1)
                                            try:
                                                driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
                                            except Exception as e2:
                                                print(e2)
                                                driver.back()

                                            time.sleep(2)
                                        elif NaviBack == "Yes" and TitleVerify == "Yes":
                                            print("Inside NaviBack=Yes TitleVerify= Yes")
                                            try:
                                                WebDriverWait(driver, SHORT_TIMEOUT
                                                              ).until(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                                WebDriverWait(driver, LONG_TIMEOUT
                                                              ).until_not(
                                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                            except TimeoutException:
                                                pass
                                            try:
                                                TitleFound = driver.find_element_by_xpath(TitleLink).text
                                                print("TitleFound is " + TitleFound)
                                                print("Title to verify is " + TitleToVerify)
                                                assert TitleFound in TitleToVerify, sheet.cell_value(ia, 1) + " not able to open"
                                                TestResult.append(sheet.cell_value(ia, 1) + " page Opened successfully")
                                                TestResultStatus.append("Pass")
                                            except Exception as e3:
                                                print(e3)
                                                TestResult.append(sheet.cell_value(ia, 1) + " page not able to open")
                                                TestResultStatus.append("Fail")

                                            time.sleep(1)
                                            try:
                                                #----------------------------------
                                                if CheckParent == 0:
                                                    print("Main module Funds clicked")
                                                    driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
                                                    time.sleep(1)
                                                    try:
                                                        driver.switch_to_alert().accept()
                                                    except Exception:
                                                        pass
                                                    try:
                                                        WebDriverWait(driver, SHORT_TIMEOUT
                                                                      ).until(
                                                            EC.presence_of_element_located(
                                                                (By.XPATH, LOADING_ELEMENT_XPATH)))

                                                        WebDriverWait(driver, LONG_TIMEOUT
                                                                      ).until_not(
                                                            EC.presence_of_element_located(
                                                                (By.XPATH, LOADING_ELEMENT_XPATH)))
                                                    except TimeoutException:
                                                        pass
                                                    # print("Browser Back clicked 1")
                                            except Exception as e2:
                                                print(e2)
                                                driver.back()
                                                # print("Browser Back clicked 2")

                                        elif NaviBack == "No" and TitleVerify == "Yes":
                                            # print("Inside NavBack no and Title Yes")
                                            for iat7 in range(1000):
                                                try:
                                                    bool = driver.find_element_by_xpath(
                                                        "//div[@id='appian-working-indicator-hidden']").is_enabled()
                                                except Exception:
                                                    time.sleep(1)
                                                    break
                                            TitleFound = driver.find_element_by_xpath(TitleLink).text
                                            # print("TitleFound1 is " + TitleFound)
                                            try:
                                                assert TitleFound in TitleToVerify, sheet.cell_value(ia, 1) + " not able to open"
                                                TestResult.append(sheet.cell_value(ia, 1) + " page Opened successfully")
                                                TestResultStatus.append("Pass")
                                            except Exception:
                                                TestResult.append(sheet.cell_value(ia, 1) + " page not able to open")
                                                TestResultStatus.append("Fail")

                                except Exception as e:
                                    print("Link not clicked / opened for  " + sheet.cell_value(ia, 1))
                                    print(e)
                                for iat4 in range(500):
                                    try:
                                        bool = driver.find_element_by_xpath(
                                            "//div[@id='appian-working-indicator-hidden']").is_enabled()
                                    except Exception:
                                        time.sleep(1)
                                        break
                except Exception as e1:
                    break
                    print(e1)
        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = ('C:/BIDS/beneficienttest/Beneficient/PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------
