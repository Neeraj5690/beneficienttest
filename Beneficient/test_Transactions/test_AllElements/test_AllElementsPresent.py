import datetime
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_AllElementsPresent"
  description = "This test scenario is to verify all the elements such as Texts, Buttons, Hyperlinks and clickable tabs are present in inside Transactions page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_Transactions"
  global Exe
  Exe="Yes"
  Directory = 'test_Transactions/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://beneficienttest.appiancloud.com/suite/")
      enter_username("neeraj.kumar")
      enter_password("Crochet@7866")
      driver.find_element_by_xpath("//input[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Ben.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_AllElementsPresent(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            print()
            PageName="Transactions"
            Ptitle1="Transactions - BIDS"
            driver.find_element_by_xpath("//*[@title='"+PageName+"']").click()
            start = time.time()
            try:
                WebDriverWait(driver, SHORT_TIMEOUT
                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                WebDriverWait(driver, LONG_TIMEOUT
                              ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            except TimeoutException:
                pass
            try:
                time.sleep(2)
                bool1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
                if bool1 == True:
                    ErrorFound1 = driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                    print(ErrorFound1)
                    driver.find_element_by_xpath(
                        "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                    TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                    TestResultStatus.append("Fail")
                    bool1 = False
                    driver.close()
            except Exception:
                try:
                    time.sleep(2)
                    bool2 = driver.find_element_by_xpath(
                        "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                    if bool2 == True:
                        ErrorFound2 = driver.find_element_by_xpath(
                            "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                        print(ErrorFound2)
                        TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                        TestResultStatus.append("Fail")
                        bool2 = False
                        driver.close()
                except Exception:
                    pass
                pass
            time.sleep(1)
            try:
                try:
                    PageTitle1 = driver.title
                    print(PageTitle1)
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                except Exception:
                    Ptitle1="Funds - BIDS"
                    PageTitle1 = driver.title
                    assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + " page Opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " page not able to open")
                TestResultStatus.append("Fail")
            stop = time.time()
            TimeString = stop - start
            print("The time of the run for " + PageName + " is: ", stop - start)
            print(TimeString)

            inside = PageName  # -----------------------------------------------------------------
            # ------Checking Transaction Listing---------
            time.sleep(2)
            Text1 = "Transaction Listing"
            Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
            try:
                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " is not present")
                TestResultStatus.append("Fail")

            # ------Checking Calendar View---------
            time.sleep(2)
            Text1 = "Calendar View"
            try:
                Element1 = driver.find_element_by_xpath(
                    "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[1]/div/div[2]/div/p/a").text
                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " inside " + inside + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " inside " + inside + " is not present")
                TestResultStatus.append("Fail")

            # ------Checking Task Management---------
            time.sleep(2)
            Text1 = "Task Management"
            try:
                Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[2]/div/div[2]/div/p/a").text

                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " inside " + inside + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " inside " + inside + " is not present")
                TestResultStatus.append("Fail")

            # ------Checking View dropdown---------
            time.sleep(2)
            Text1 = "View:"
            try:
                Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[3]/div/div[1]/span").text

                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " inside " + inside + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " inside " + inside + " is not present")
                TestResultStatus.append("Fail")

            # ------Checking Filter By: dropdown---------
            time.sleep(2)
            Text1 = "Filter By:"
            try:
                Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[4]/div/div[1]/span").text

                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " inside " + inside + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " inside " + inside + " is not present")
                TestResultStatus.append("Fail")

            # ------Checking Project Name---------
            time.sleep(2)
            Text1 = "Project Name"
            try:
                Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[5]/div/div[1]/label").text

                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " inside " + inside + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " inside " + inside + " is not present")
                TestResultStatus.append("Fail")

            # # ------Checking Last Updated On---------
            # time.sleep(2)
            # Text1 = "Last Updated On"
            # try:
            #     Element1 = driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[6]/div/div[1]/span").text
            #     assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
            #     TestResult.append(Text1 + " inside " + inside + " is present")
            #     TestResultStatus.append("Pass")
            # except Exception as e1:
            #     print(e1)
            #     TestResult.append(Text1 + " inside " + inside + " is not present")
            #     TestResultStatus.append("Fail")

            # ------Checking Reload Data---------
            time.sleep(2)
            Text1 = "Reload Data"
            try:
                Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[6]/div/div[2]/div/p/a").text
                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " inside " + inside + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " inside " + inside + " is not present")
                TestResultStatus.append("Fail")

            # ------Checking Reload Data---------
            time.sleep(2)
            Text1 = "EXPORT TO EXCEL"
            try:
                Element1 = driver.find_element_by_xpath(
                "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[7]/div/div/button").text
                assert Text1 in Element1, Text1 + " inside " + inside + " is not present"
                TestResult.append(Text1 + " inside " + inside + " is present")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(Text1 + " inside " + inside + " is not present")
                TestResultStatus.append("Fail")

            # ---------------loop for Columns in table for Transaction Listing----------
            ItemList = ["Project Name", "Transaction ID", "Date of Intro to Deal", "Status", "Stage Lost (if applicable)", "Seller", "Intermediary", "Process Type", "Channel / Seller Type", "Opportunity Type", "Asset Opportunity","# of Funds (submitted)","# of Investments (submitted)","NAV (submitted)","Assigned U/W Deal Leader","Asset / GP Commentary" ]
            for ii in range(len(ItemList)):
                print(str(ii + 1))
                Text1 = ItemList[ii]
                print("Text1 " + Text1)
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]//table/thead/tr[1]/th[" + str(
                            ii + 1) + "]/div").text
                    #print("Element1 " + Element1)
                except Exception as ee:
                    #print(ee)
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column is not present in Transaction Listing table"
                    TestResult.append(
                        Text1 + " column is present in Transaction Listing table")
                    TestResultStatus.append("Pass")
                except Exception as e1:
                    print(e1)
                    TestResult.append(
                        Text1 + " column is not present in Transaction Listing table")
                    TestResultStatus.append("Fail")


            # #---------------------------------------------------------------
            # PageName = "Calendar View"
            # driver.find_element_by_xpath("//*[text() = '" + PageName + "']").click()
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # try:
            #     time.sleep(2)
            #     bool1 = driver.find_element_by_xpath(
            #         "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
            #     if bool1 == True:
            #         ErrorFound1 = driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
            #         print(ErrorFound1)
            #         driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
            #         TestResult.append(PageName + " not able to open\n" + ErrorFound1)
            #         TestResultStatus.append("Fail")
            #         bool1 = False
            # except Exception:
            #     try:
            #         time.sleep(2)
            #         bool2 = driver.find_element_by_xpath(
            #             "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
            #         if bool2 == True:
            #             ErrorFound2 = driver.find_element_by_xpath(
            #                 "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
            #             print(ErrorFound2)
            #             TestResult.append(PageName + " not able to open\n" + ErrorFound2)
            #             TestResultStatus.append("Fail")
            #             bool2 = False
            #     except Exception:
            #         pass
            #     pass
            # time.sleep(1)
            # Ptitle2 = "Transaction Listing"
            # PageTitle2 = driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
            # try:
            #     print(PageTitle2)
            #     assert Ptitle2 in PageTitle2, PageName + " not able to open"
            #     TestResult.append(PageName + " Opened successfully")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " not able to open")
            #     TestResultStatus.append("Fail")
            # driver.find_element_by_xpath("//*[@title='Transactions']").click()
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # time.sleep(1)
            #
            # inside="Calendar View" #-----------------------------------------------------------------
            # # ------Checking Transaction Listing---------
            # time.sleep(2)
            # Text1 = "Transaction Listing"
            # Element1 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[1]/div/div/div").text
            # try:
            #     assert Text1 in Element1, Text1+" inside "+inside+" is not present"
            #     TestResult.append(Text1+" is present")
            #     TestResultStatus.append("Pass")
            # except Exception as e1:
            #     print(e1)
            #     TestResult.append(Text1+" is not present")
            #     TestResultStatus.append("Fail")
            #
            #
            #
            #
            #
            # #----------------------------------------------------------------
            # PageName = "Task Management"
            # driver.find_element_by_xpath("//*[text() = '"+PageName+"']").click()
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # try:
            #     time.sleep(2)
            #     bool1 = driver.find_element_by_xpath(
            #         "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
            #     if bool1 == True:
            #         ErrorFound1 = driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
            #         print(ErrorFound1)
            #         driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
            #         TestResult.append(PageName + " not able to open\n" + ErrorFound1)
            #         TestResultStatus.append("Fail")
            #         bool1 = False
            # except Exception:
            #     try:
            #         time.sleep(2)
            #         bool2 = driver.find_element_by_xpath(
            #             "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
            #         if bool2 == True:
            #             ErrorFound2 = driver.find_element_by_xpath(
            #                 "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
            #             print(ErrorFound2)
            #             TestResult.append(PageName + " not able to open\n" + ErrorFound2)
            #             TestResultStatus.append("Fail")
            #             bool2 = False
            #     except Exception:
            #         pass
            #     pass
            # time.sleep(1)
            # Ptitle2 = "Task Management"
            # PageTitle2 = driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div/div[1]/div").text
            # try:
            #     print(PageTitle2)
            #     assert Ptitle2 in PageTitle2, PageName + " not able to open"
            #     TestResult.append(PageName + " page Opened successfully")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " page not able to open")
            #     TestResultStatus.append("Fail")
            # driver.find_element_by_xpath("//*[@title='Transactions']").click()
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # time.sleep(1)
            #
            # #----------------------------------------------------------------------
            # PageName = "Transaction ID"
            # try:
            #     driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr[1]/td[2]/div/p/a").click()
            #     start = time.time()
            # except Exception:
            #     time.sleep(7)
            #     try:
            #         driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div[2]/div/div[1]/div[2]/table/tbody/tr[1]/td[2]/div/p/a").click()
            #     except Exception:
            #         TestResult.append(PageName + " not able to open on click")
            #         TestResultStatus.append("Fail")
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # try:
            #     time.sleep(2)
            #     bool1 = driver.find_element_by_xpath(
            #         "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
            #     if bool1 == True:
            #         ErrorFound1 = driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
            #         print(ErrorFound1)
            #         driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
            #         TestResult.append(PageName + " not able to open\n" + ErrorFound1)
            #         TestResultStatus.append("Fail")
            #         bool1 = False
            # except Exception:
            #     try:
            #         time.sleep(2)
            #         bool2 = driver.find_element_by_xpath(
            #             "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
            #         if bool2 == True:
            #             ErrorFound2 = driver.find_element_by_xpath(
            #                 "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
            #             print(ErrorFound2)
            #             TestResult.append(PageName + " not able to open\n" + ErrorFound2)
            #             TestResultStatus.append("Fail")
            #             bool2 = False
            #     except Exception:
            #         pass
            #     pass
            # time.sleep(1)
            # Ptitle3 = "Transaction NAV Concentration"
            # PageTitle3 = driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[2]/div/div/div/div/div[1]/div[1]/div[1]/div/div/div").text
            # try:
            #     assert Ptitle3 in PageTitle3, PageName + " not able to open"
            #     TestResult.append(PageName + " clicked and opened successfully")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " not able to open on click")
            #     TestResultStatus.append("Fail")
            # stop = time.time()
            # TimeString = stop - start
            # print("The time of the run for " + PageName + " is: ", stop - start)
            # print(TimeString)
            #
            # #------------------------------------------------------------
            # PageName = "DEAL TEMPLATE"
            # driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[3]/button").click()
            # start = time.time()
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # try:
            #     time.sleep(2)
            #     bool1 = driver.find_element_by_xpath(
            #         "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
            #     if bool1 == True:
            #         ErrorFound1 = driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
            #         print(ErrorFound1)
            #         driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
            #         TestResult.append(PageName + " not able to open\n" + ErrorFound1)
            #         TestResultStatus.append("Fail")
            #         bool1 = False
            # except Exception:
            #     try:
            #         time.sleep(2)
            #         bool2 = driver.find_element_by_xpath(
            #             "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
            #         if bool2 == True:
            #             ErrorFound2 = driver.find_element_by_xpath(
            #                 "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
            #             print(ErrorFound2)
            #             TestResult.append(PageName + " not able to open\n" + ErrorFound2)
            #             TestResultStatus.append("Fail")
            #             bool2 = False
            #     except Exception:
            #         pass
            #     pass
            # time.sleep(1)
            # Ptitle4 = "Deal Template Status Summary"
            # PageTitle4 = driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[4]/div[1]/div/div[1]/div[1]/div[1]/div/div[1]/div").text
            # try:
            #     print(PageTitle3)
            #     assert Ptitle4 in PageTitle4, PageName + " not able to open"
            #     TestResult.append(PageName + " button clicked successfully")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " button not able to open on click")
            #     TestResultStatus.append("Fail")
            # stop = time.time()
            # TimeString = stop - start
            # print("The time of the run for " + PageName + " is: ", stop - start)
            # print(TimeString)
            #
            # # ------------------------------------------------------------
            # PageName = "FUND DETAILS"
            # driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[1]/button").click()
            # start = time.time()
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # try:
            #     time.sleep(2)
            #     bool1 = driver.find_element_by_xpath(
            #         "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
            #     if bool1 == True:
            #         ErrorFound1 = driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
            #         print(ErrorFound1)
            #         driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
            #         TestResult.append(PageName + " not able to open\n" + ErrorFound1)
            #         TestResultStatus.append("Fail")
            #         bool1 = False
            # except Exception:
            #     try:
            #         time.sleep(2)
            #         bool2 = driver.find_element_by_xpath(
            #             "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
            #         if bool2 == True:
            #             ErrorFound2 = driver.find_element_by_xpath(
            #                 "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
            #             print(ErrorFound2)
            #             TestResult.append(PageName + " not able to open\n" + ErrorFound2)
            #             TestResultStatus.append("Fail")
            #             bool2 = False
            #     except Exception:
            #         pass
            #     pass
            # time.sleep(1)
            # Ptitle4 = "View By"
            # PageTitle4 = driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[4]/div[1]/div/div[1]/span").text
            # try:
            #     print(PageTitle3)
            #     assert Ptitle4 in PageTitle4, PageName + " not able to open"
            #     TestResult.append(PageName + " button clicked successfully")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " button not able to open on click")
            #     TestResultStatus.append("Fail")
            # stop = time.time()
            # TimeString = stop - start
            # print("The time of the run for " + PageName + " is: ", stop - start)
            # print(TimeString)
            #
            # #-------------------------------------------------------------
            # PageName = "Transaction Mgmt"
            # driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[3]/div/div/div[2]/button").click()
            # start = time.time()
            # try:
            #     WebDriverWait(driver, SHORT_TIMEOUT
            #                   ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #     WebDriverWait(driver, LONG_TIMEOUT
            #                   ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            # except TimeoutException:
            #     pass
            # try:
            #     time.sleep(2)
            #     bool1 = driver.find_element_by_xpath(
            #         "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
            #     if bool1 == True:
            #         ErrorFound1 = driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
            #         print(ErrorFound1)
            #         driver.find_element_by_xpath(
            #             "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
            #         TestResult.append(PageName + " not able to open\n" + ErrorFound1)
            #         TestResultStatus.append("Fail")
            #         bool1 = False
            #         driver.close()
            # except Exception:
            #     try:
            #         time.sleep(2)
            #         bool2 = driver.find_element_by_xpath(
            #             "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
            #         if bool2 == True:
            #             ErrorFound2 = driver.find_element_by_xpath(
            #                 "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
            #             print(ErrorFound2)
            #             TestResult.append(PageName + " not able to open\n" + ErrorFound2)
            #             TestResultStatus.append("Fail")
            #             bool2 = False
            #             driver.close()
            #     except Exception:
            #         pass
            #     pass
            # time.sleep(1)
            # Ptitle4 = "Transaction Workflow"
            # PageTitle4 = driver.find_element_by_xpath(
            #     "//div[@class='ContentLayout---content_layout']/div[4]/div[1]/div/div[1]/div[1]").text
            # try:
            #     print(PageTitle3)
            #     assert Ptitle4 in PageTitle4, PageName + " not able to open"
            #     TestResult.append(PageName + " button clicked successfully")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " button not able to open on click")
            #     TestResultStatus.append("Fail")
            # stop = time.time()
            # TimeString = stop - start
            # print("The time of the run for " + PageName + " is: ", stop - start)
            # print(TimeString)

        except Exception as Mainerror:
            stop = time.time()
            RoundFloatString = round(float(stop - start),2)
            print("The time of the run for " + PageName + " is: ", RoundFloatString)
            stringMainerror=repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


